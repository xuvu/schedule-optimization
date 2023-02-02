"""Nurse scheduling problem with shift requests."""
import math
from datetime import date

from ortools.sat.python import cp_model
import calendar
import numpy as np
from pandas import date_range
import itertools

name_of_person = ['ภูวเนตร',
                  'ราเชนทร์',
                  'พลกฤต',
                  'ชานนท์',
                  'ปรมะ',
                  'สัญญา',
                  'วินัย',
                  'รณยุทธ',
                  'วัฒพงษ์',
                  'ราเชน',
                  'นฤชิต',
                  'จีรวัฒน์',
                  'สุเมธร์',
                  'นที'
                  ]

# Type of nurses
# programmer = 8, service = 6
roles = ["Programmer", "Programmer", "Programmer", "Programmer", "Programmer", "Programmer", "Programmer", "Programmer",
         "Service", "Service", "Service", "Service", "Service", "Service"]


# Get amount of each type
def get_amount_of_each_type(r):
    counts = {}
    for role in r:
        if role in counts:
            counts[role] += 1
        else:
            counts[role] = 1

    return counts


# Get the name of a nurse
def map_name_person(n):
    return name_of_person[n]


def get_weekend_days_in_year(year):
    weekend_days = []
    normal_day = 1
    for month in range(1, 13):
        for day in range(1, calendar.monthrange(year, month)[1] + 1):
            current_date = date(year, month, day)
            if current_date.weekday() in [5, 6]:
                weekend_days.append(normal_day)
            normal_day += 1
    return weekend_days


def get_days_in_year(year):
    days_in_year = [i for i in range(1, 366)]
    return days_in_year


def create_list(number_of_nurse, days_in_month, num_shift_per_day, num_nurse_per_shift):
    # create list which has shape like (number_of_nurse,days_in_month ,3)
    # 3 is for 3 shifts
    # 0 is for day off
    # 1 is for day on
    # 2 is for weekend
    return np.zeros((number_of_nurse, days_in_month, num_shift_per_day, num_nurse_per_shift))


def is_holiday(day, list_of_holiday):
    if day in list_of_holiday:
        return True
    else:
        return False


def get_max_diff_type(r, total_shifts):
    role_count = get_amount_of_each_type(r)

    min_value = min(role_count.values())
    min_type = [key for key, value in role_count.items() if value == min_value]

    max_value = max(role_count.values())
    max_type = [key for key, value in role_count.items() if value == max_value]

    list1 = [f"{k}{i}" for i, k in enumerate(r, 1) if k == str(max_type[0])]
    list2 = [f"{k}{i}" for i, k in enumerate(r, 1) if k == str(min_type[0])]

    min_shift = math.floor(total_shifts / (len(list1) + len(list2)))

    element_counts = {element: 0 for element in list1 + list2}
    result = []

    count_pair = 0
    for k in range(math.floor(total_shifts / (len(list1)) * len(list2))):
        for b in list2:
            for a in list1:
                if count_pair >= total_shifts:
                    break
                if b != a and element_counts[b] < min_shift and element_counts[a] < min_shift:
                    count_pair += 1
                    element_counts[a] += 1
                    element_counts[b] += 1
                    result += [(a, b)]
    return count_pair


# Define the types of nurses
types = ["Programmer", "Service"]
all_types = range(len(types))

# (EASY) (FINISH)
# no morning shift in working day => good

# (EASY) (FINISH)
# holiday can have all three shifts => good

# (FINISH)
# Each shift has 2 person

# (MEDIUM) (FINISH)
# 2 person per shift
# programmer + service => programmer + programmer => service + service

# (HARD)
# (1,2) same => (1,2) different => (0,1) same => good

# (HARD)
# night to morning is forbidden (0,2) in a day (can but shouldn't) => good

# (HARD)
# morning to night is forbidden (2,0) in a day (forbidden) => good

# special holiday

# special hospital
# afternoon to night (1,2) the highest priority with same person => (1,2) different person
# (0,1) if (1,2) can't happen

year_ = 2023
month_ = 1


# Get type index of a nurse
def type_of_nurse(n):
    return types.index(roles[n])


# Get name of type of nurse
def name_of_type_nurse(n):
    return roles[n]


def main():
    # This program tries to find an optimal assignment of nurses to shifts
    # (3 shifts per day, for 7 days), subject to some constraints (see below).
    # Each nurse can request to be assigned to specific shifts.
    # The optimal assignment maximizes the number of fulfilled shift requests

    global total_min, total_max
    num_nurses = len(roles)
    all_nurses = range(num_nurses)
    all_days = get_days_in_year(year_)

    all_weekends = get_weekend_days_in_year(year_)
    all_working_days = set(all_days) - set(all_weekends)

    total_working_days = len(all_working_days)
    total_holidays = len(all_weekends)

    # shift_requests[0][0][0] means first nurse, first day, first shift shift_requests[1][2][1] would mean that the
    # second nurse wants to work on the third day of the week on the second shift of the day.

    # Creates the model.
    model = cp_model.CpModel()

    # Creates shift variables.
    # Shifts[(n, d, s)]: nurse 'n' works shift 's' on day 'd'.
    shifts = {}
    num_shifts = 3
    all_shifts = range(num_shifts)  # [1,2,3,4]

    forbidden_shifts = [0]

    # Number of nurse per shift
    num_of_nurse_per_shift = 3
    all_nurse_per_shift = range(num_of_nurse_per_shift)

    # this variable should be generated automatically then can be modified later on
    shift_requests = create_list(num_nurses, len(all_days), num_shifts, num_of_nurse_per_shift)  # <-- demo

    for n in all_nurses:
        for d in all_days:
            for s in all_shifts:
                for k in all_nurse_per_shift:
                    shifts[(n, d, s, k)] = model.NewBoolVar('shift_n%id%is%ik%i' % (n, d, s, k))

    """
    # Each shift is assigned to exactly one nurse in.
    for d in all_days:
        for s in all_shifts:
            model.Add(sum(shifts[(n, d, s)] for n in all_nurses) == 1)
    """
    # forbidden morning shift in working day
    for n in all_nurses:
        for d in all_working_days:
            for s in all_shifts:
                if s in forbidden_shifts:
                    for k in all_nurse_per_shift:
                        model.Add(shifts[(n, d, s, k)] == 0)

    # Each shift is assigned to exactly 1 nurse in.
    for d in all_days:
        if d not in all_weekends:  # not in weekend => in working day
            for s in all_shifts:
                for k in all_nurse_per_shift:
                    if s not in forbidden_shifts:  # exclude morning shift
                        model.Add(sum(shifts[(n, d, s, k)] for n in all_nurses) == 1)
        else:
            for s in all_shifts:
                for k in all_nurse_per_shift:
                    model.Add(sum(shifts[(n, d, s, k)] for n in all_nurses) == 1)

    # For each nurse shouldn't have more than 2 shifts within 4 days
    day_interval = 4
    maximum_shifts = 2
    for n in all_nurses:
        for d in all_days:  # d = 1,2,3,4,5,...,total_day, d = 1
            if (d + day_interval) < len(all_days):  # 1 + (4-1) < 31
                model.Add(sum(shifts[(n, d + i, s, k)] for i in range(day_interval) for s in all_shifts for k in
                              all_nurse_per_shift) <= maximum_shifts)
                # shifts[(0, 1 + 0, 0, k), shifts[(0, 1 + 0, 1, k), shifts[(0, 1 + 0, 2, k)
                # shifts[(0, 1 + 1, 0, k), shifts[(0, 1 + 1, 1, k), shifts[(0, 1 + 1, 2, k)
                # shifts[(0, 1 + 2, 0, k), shifts[(0, 1 + 2, 1, k), shifts[(0, 1 + 2, 2, k)
                # shifts[(0, 1 + 3, 0, k), shifts[(0, 1 + 3, 1, k), shifts[(0, 1 + 3, 2, k)
                # all of the shifts period the nurse can only have 2 shifts maximum

    # Each nurse works at most two shifts per day.
    maximum_each_day_shifts = 2
    for n in all_nurses:
        for d in all_days:
            shifts_per_day = sum(shifts[(n, d, s, k)] for s in all_shifts for k in all_nurse_per_shift)
            model.Add(shifts_per_day <= maximum_each_day_shifts)

    # Nurse type priority
    type_priority = [['Type1', 'Type2'], ['Type1', 'Type1'], ['Type2', 'Type2']]

    # total possible way of pairing ['Type1', 'Type2'] is the product of the number of people for each type

    # Prioritize shift patterns
    for d in all_days:
        for n in all_nurses:
            for k in all_nurse_per_shift:
                model.Add(shifts[(n, d, 1, k)] <= shifts[(n, d, 2, k)])  # (1,2) same

    '''
    # Prioritize shift patterns
    for d in all_days:
        for n in all_nurses:
            if d <= 20:
                model.Add(shifts[(n, d, 1)] <= shifts[(n, d, 2)])  # (1,2) same
            else:
                model.Add(shifts[(n, d, 0)] <= shifts[(n, d, 1)])  # (0,1) same

            model.Add(shifts[(n, d, 0)] + shifts[(n, d, 2)] <= 1)  # forbidden (0,2)
            model.Add(shifts[(n, d, 2)] + shifts[(n, d, 0)] <= 1)  # forbidden (2,0)
    '''
    # Try to distribute the shifts evenly, so that each nurse works
    # min_shifts_per_nurse shifts. If this is not possible, because the total
    # number of shifts is not divisible by the number of nurses, some nurses will
    # be assigned one more shift.

    # min-max working day
    working_shift_per_day = ((num_shifts - len(forbidden_shifts)) * (num_of_nurse_per_shift - 1))
    min_shifts_per_nurse = (working_shift_per_day * total_working_days) // num_nurses
    if working_shift_per_day * total_working_days % num_nurses == 0:
        max_shifts_per_nurse = min_shifts_per_nurse
    else:
        max_shifts_per_nurse = min_shifts_per_nurse + 1

    # min-max holiday day
    min_shifts_per_nurse_h = ((num_shifts * (num_of_nurse_per_shift - 1)) * total_holidays) // num_nurses
    if (num_shifts * (num_of_nurse_per_shift - 1)) * total_holidays % num_nurses == 0:
        max_shifts_per_nurse_h = min_shifts_per_nurse_h
    else:
        max_shifts_per_nurse_h = min_shifts_per_nurse_h + 1

    # min-max working day (S)
    working_shift_per_day = ((num_shifts - len(forbidden_shifts)) * (num_of_nurse_per_shift - 2))
    min_shifts_per_nurse_S = (working_shift_per_day * total_working_days) // num_nurses
    if working_shift_per_day * total_working_days % num_nurses == 0:
        max_shifts_per_nurse_S = min_shifts_per_nurse_S
    else:
        max_shifts_per_nurse_S = min_shifts_per_nurse_S + 1

    # min-max holiday day (S)
    min_shifts_per_nurse_h_S = ((num_shifts * (num_of_nurse_per_shift - 2)) * total_holidays) // num_nurses
    if (num_shifts * (num_of_nurse_per_shift - 2)) * total_holidays % num_nurses == 0:
        max_shifts_per_nurse_h_S = min_shifts_per_nurse_h_S
    else:
        max_shifts_per_nurse_h_S = min_shifts_per_nurse_h_S + 1

    # All working shifts min-max variable
    # ---------------------------------------------------------------
    min_morning_shift_per_nurse = (total_working_days * (num_of_nurse_per_shift - 1)) // num_nurses
    min_afternoon_shift_per_nurse = (total_working_days * (num_of_nurse_per_shift - 1)) // num_nurses
    min_night_shift_per_nurse = (total_working_days * (num_of_nurse_per_shift - 1)) // num_nurses

    min_morning_shift_per_nurse_S = (total_working_days * (num_of_nurse_per_shift - 2)) // num_nurses
    min_afternoon_shift_per_nurse_S = (total_working_days * (num_of_nurse_per_shift - 2)) // num_nurses
    min_night_shift_per_nurse_S = (total_working_days * (num_of_nurse_per_shift - 2)) // num_nurses

    if min_morning_shift_per_nurse_S == 0 and (total_working_days * (num_of_nurse_per_shift - 2)) != 0:
        min_morning_shift_per_nurse_S = 1
        max_morning_shift_per_nurse_S = 1

        min_afternoon_shift_per_nurse_S = 1
        max_afternoon_shift_per_nurse_S = 1

        min_night_shift_per_nurse_S = 1
        max_night_shift_per_nurse_S = 1

    elif (total_working_days * (num_of_nurse_per_shift - 2)) % num_nurses == 0:
        max_morning_shift_per_nurse_S = min_morning_shift_per_nurse_S
        max_afternoon_shift_per_nurse_S = min_morning_shift_per_nurse_S
        max_night_shift_per_nurse_S = min_morning_shift_per_nurse_S
    else:
        max_morning_shift_per_nurse_S = min_morning_shift_per_nurse_S + 1
        max_afternoon_shift_per_nurse_S = min_morning_shift_per_nurse_S + 1
        max_night_shift_per_nurse_S = min_morning_shift_per_nurse_S + 1

    if min_morning_shift_per_nurse == 0 and (total_working_days * (num_of_nurse_per_shift - 1)) != 0:
        min_morning_shift_per_nurse = 1
        max_morning_shift_per_nurse = 1

        min_afternoon_shift_per_nurse = 1
        max_afternoon_shift_per_nurse = 1

        min_night_shift_per_nurse = 1
        max_night_shift_per_nurse = 1

    elif (total_working_days * (num_of_nurse_per_shift - 1)) % num_nurses == 0:
        max_morning_shift_per_nurse = min_morning_shift_per_nurse
        max_afternoon_shift_per_nurse = min_morning_shift_per_nurse
        max_night_shift_per_nurse = min_morning_shift_per_nurse
    else:
        max_morning_shift_per_nurse = min_morning_shift_per_nurse + 1
        max_afternoon_shift_per_nurse = min_morning_shift_per_nurse + 1
        max_night_shift_per_nurse = min_morning_shift_per_nurse + 1
    # ---------------------------------------------------------------

    # All holiday shifts min-max variable
    # ---------------------------------------------------------------
    min_morning_shift_per_nurse_h = (total_holidays * (num_of_nurse_per_shift - 1)) // num_nurses
    min_afternoon_shift_per_nurse_h = (total_holidays * (num_of_nurse_per_shift - 1)) // num_nurses
    min_night_shift_per_nurse_h = (total_holidays * (num_of_nurse_per_shift - 1)) // num_nurses

    min_morning_shift_per_nurse_h_S = (total_holidays * (num_of_nurse_per_shift - 2)) // num_nurses
    min_afternoon_shift_per_nurse_h_S = (total_holidays * (num_of_nurse_per_shift - 2)) // num_nurses
    min_night_shift_per_nurse_h_S = (total_holidays * (num_of_nurse_per_shift - 2)) // num_nurses

    if min_morning_shift_per_nurse_h_S == 0 and (total_holidays * (num_of_nurse_per_shift - 2)) != 0:
        min_morning_shift_per_nurse_h_S = 1
        max_morning_shift_per_nurse_h_S = 1

        min_afternoon_shift_per_nurse_h_S = 1
        max_afternoon_shift_per_nurse_h_S = 1

        min_night_shift_per_nurse_h_S = 1
        max_night_shift_per_nurse_h_S = 1

    elif (total_holidays * (num_of_nurse_per_shift - 2)) % num_nurses == 0:
        max_morning_shift_per_nurse_h_S = min_morning_shift_per_nurse_h_S
        max_afternoon_shift_per_nurse_h_S = min_afternoon_shift_per_nurse_h_S
        max_night_shift_per_nurse_h_S = min_night_shift_per_nurse_h_S

    else:
        max_morning_shift_per_nurse_h_S = min_morning_shift_per_nurse_h_S + 1
        max_afternoon_shift_per_nurse_h_S = min_afternoon_shift_per_nurse_h_S + 1
        max_night_shift_per_nurse_h_S = min_night_shift_per_nurse_h_S + 1

    if min_morning_shift_per_nurse_h == 0 and (total_holidays * (num_of_nurse_per_shift - 1)) != 0:
        min_morning_shift_per_nurse_h = 1
        max_morning_shift_per_nurse_h = 1

        min_afternoon_shift_per_nurse_h = 1
        max_afternoon_shift_per_nurse_h = 1

        min_night_shift_per_nurse_h = 1
        max_night_shift_per_nurse_h = 1

    elif (total_holidays * (num_of_nurse_per_shift - 1)) % num_nurses == 0:
        max_morning_shift_per_nurse_h = min_morning_shift_per_nurse_h
        max_afternoon_shift_per_nurse_h = min_afternoon_shift_per_nurse_h
        max_night_shift_per_nurse_h = min_night_shift_per_nurse_h
    else:
        max_morning_shift_per_nurse_h = min_morning_shift_per_nurse_h + 1
        max_afternoon_shift_per_nurse_h = min_afternoon_shift_per_nurse_h + 1
        max_night_shift_per_nurse_h = min_night_shift_per_nurse_h + 1
    # ---------------------------------------------------------------

    # Distribute all day for main hospital
    for n in all_nurses:
        num_shift_worked = 0
        num_shift_worked_h = 0

        num_shifts_morning = 0
        num_shifts_morning_h = 0

        num_shifts_afternoon = 0
        num_shifts_afternoon_h = 0

        num_shifts_night = 0
        num_shifts_night_h = 0

        num_shift_worked_S = 0
        num_shift_worked_h_S = 0

        num_shifts_morning_S = 0
        num_shifts_morning_h_S = 0

        num_shifts_afternoon_S = 0
        num_shifts_afternoon_h_S = 0

        num_shifts_night_S = 0
        num_shifts_night_h_S = 0

        for k in all_nurse_per_shift:
            if k != 2:
                # Distribute working day shifts
                # ---------------------------------------------------------------
                for d in all_working_days:
                    for s in all_shifts:
                        if s == 1:
                            num_shift_worked += shifts[(n, d, s, k)]
                            num_shifts_afternoon += shifts[(n, d, s, k)]

                        elif s == 2:
                            num_shift_worked += shifts[(n, d, s, k)]
                            num_shifts_night += shifts[(n, d, s, k)]

                # Distribute holiday day shifts
                # ---------------------------------------------------------------
                for d in all_weekends:
                    for s in all_shifts:
                        if s == 0:
                            num_shift_worked_h += shifts[(n, d, s, k)]
                            num_shifts_morning_h += shifts[(n, d, s, k)]
                        elif s == 1:
                            num_shift_worked_h += shifts[(n, d, s, k)]
                            num_shifts_afternoon_h += shifts[(n, d, s, k)]
                        elif s == 2:
                            num_shift_worked_h += shifts[(n, d, s, k)]
                            num_shifts_night_h += shifts[(n, d, s, k)]
            else:
                # Distribute working day shifts (S)
                # ---------------------------------------------------------------
                for d in all_working_days:
                    for s in all_shifts:
                        if s == 1:
                            num_shift_worked_S += shifts[(n, d, s, k)]
                            num_shifts_afternoon_S += shifts[(n, d, s, k)]

                        elif s == 2:
                            num_shift_worked_S += shifts[(n, d, s, k)]
                            num_shifts_night_S += shifts[(n, d, s, k)]

                # Distribute holiday day shifts (S)
                # ---------------------------------------------------------------
                for d in all_weekends:
                    for s in all_shifts:
                        if s == 0:
                            num_shift_worked_h_S += shifts[(n, d, s, k)]
                            num_shifts_morning_h_S += shifts[(n, d, s, k)]
                        elif s == 1:
                            num_shift_worked_h_S += shifts[(n, d, s, k)]
                            num_shifts_afternoon_h_S += shifts[(n, d, s, k)]
                        elif s == 2:
                            num_shift_worked_h_S += shifts[(n, d, s, k)]
                            num_shifts_night_h_S += shifts[(n, d, s, k)]

        # all shifts constraint working day
        # model.Add(min_shifts_per_nurse <= num_shift_worked)
        # model.Add(num_shift_worked <= max_shifts_per_nurse)
        #
        # # all shifts constraint holiday
        # model.Add(min_shifts_per_nurse_h <= num_shift_worked_h)
        # model.Add(num_shift_worked_h <= max_shifts_per_nurse_h)
        #
        # # all shifts constraint working day (S)
        # model.Add(min_shifts_per_nurse_S <= num_shift_worked_S)
        # model.Add(num_shift_worked_S <= max_shifts_per_nurse_S)
        #
        # # all shifts constraint holiday (S)
        # model.Add(min_shifts_per_nurse_h_S <= num_shift_worked_h_S)
        # model.Add(num_shift_worked_h_S <= max_shifts_per_nurse_h_S)

        # morning shifts constraint (Forbidden)
        # model.Add(min_morning_shift_per_nurse <= num_shifts_morning)
        # model.Add(num_shifts_morning <= max_morning_shift_per_nurse)

        # conflict between
        # minimum type within each shift
        # distribution constraint

        # soften the constraint of working day shifts
        slack = 0

        # afternoon shifts constraint working day
        model.Add(min_afternoon_shift_per_nurse - slack <= num_shifts_afternoon)
        model.Add(num_shifts_afternoon <= max_afternoon_shift_per_nurse + slack)

        # night shifts constraint working day
        model.Add(min_night_shift_per_nurse - slack <= num_shifts_night)
        model.Add(num_shifts_night <= max_night_shift_per_nurse + slack)

        # afternoon shifts constraint working day (S)
        model.Add(min_afternoon_shift_per_nurse_S - slack <= num_shifts_afternoon_S)
        model.Add(num_shifts_afternoon_S <= max_afternoon_shift_per_nurse_S + slack)

        # night shifts constraint working day (S)
        model.Add(min_night_shift_per_nurse_S - slack <= num_shifts_night_S)
        model.Add(num_shifts_night_S <= max_night_shift_per_nurse_S + slack)

        # soften the constraint for holiday shifts
        slack_h = 0

        # morning shifts constraint holiday
        model.Add(min_morning_shift_per_nurse_h - slack_h <= num_shifts_morning_h)
        model.Add(num_shifts_morning_h <= max_morning_shift_per_nurse_h + slack_h)

        # afternoon shifts constraint holiday
        model.Add(min_afternoon_shift_per_nurse_h - slack_h <= num_shifts_afternoon_h)
        model.Add(num_shifts_afternoon_h <= max_afternoon_shift_per_nurse_h + slack_h)

        # night shifts constraint holiday
        model.Add(min_night_shift_per_nurse_h - slack_h <= num_shifts_night_h)
        model.Add(num_shifts_night_h <= max_night_shift_per_nurse_h + slack_h)

        # morning shifts constraint holiday (S)
        model.Add(min_morning_shift_per_nurse_h_S - slack_h <= num_shifts_morning_h_S)
        model.Add(num_shifts_morning_h_S <= max_morning_shift_per_nurse_h_S + slack_h)

        # afternoon shifts constraint holiday (S)
        model.Add(min_afternoon_shift_per_nurse_h_S - slack_h <= num_shifts_afternoon_h_S)
        model.Add(num_shifts_afternoon_h_S <= max_afternoon_shift_per_nurse_h_S + slack_h)

        # night shifts constraint holiday (S)
        model.Add(min_night_shift_per_nurse_h_S - slack_h <= num_shifts_night_h_S)
        model.Add(num_shifts_night_h_S <= max_night_shift_per_nurse_h_S + slack_h)

        # lower bound slack
        min_total_slack = 0

        # upper bound slack
        max_total_slack = 0

        total_min = min_shifts_per_nurse + min_shifts_per_nurse_h + min_shifts_per_nurse_S + min_shifts_per_nurse_h_S - min_total_slack
        total_max = max_shifts_per_nurse + max_shifts_per_nurse_h + max_shifts_per_nurse_S + max_shifts_per_nurse_h_S + max_total_slack

        # Total shifts should be equal
        # model.Add(total_min <= (num_shift_worked + num_shift_worked_h + num_shift_worked_S + num_shift_worked_h_S))
        # model.Add((num_shift_worked + num_shift_worked_h + num_shift_worked_S + num_shift_worked_h_S) <= total_max)
        # ---------------------------------------------------------------

    all_shift_count = ((total_holidays * (num_shifts - 0)) + (
            total_working_days * (num_shifts - len(forbidden_shifts)))) * num_of_nurse_per_shift
    max_diff_pair = get_max_diff_type(roles, all_shift_count)
    max_diff_count = 0

    # Each shift ensure different type of nurses
    # for d in all_weekends:
    #     for s in all_shifts:
    #         if max_diff_count <= max_diff_pair:
    #             max_diff_count += 1
    #             type_count = [0] * len(all_types)
    #             for k in range(0, num_of_nurse_per_shift - 1):  # -1 for excluding the third person
    #                 for t in all_types:
    #                     for n in all_nurses:
    #                         if type_of_nurse(n) == t:
    #                             type_count[t] += shifts[(n, d, s, k)]
    #             for t in all_types:
    #                 model.Add(type_count[t] <= 1)
    #
    # for d in all_working_days:
    #     for s in all_shifts:
    #         if s != 0:  # exclude morning and special hospital
    #             if max_diff_count <= max_diff_pair:
    #                 max_diff_count += 1
    #                 type_count = [0] * len(all_types)
    #                 for k in range(0, num_of_nurse_per_shift - 1):  # -1 for excluding the third person
    #                     for t in all_types:
    #                         for n in all_nurses:
    #                             if type_of_nurse(n) == t:
    #                                 type_count[t] += shifts[(n, d, s, k)]
    #                 for t in all_types:
    #                     model.Add(type_count[t] <= 1)

    # pylint: disable=g-complex-comprehension
    model.Maximize(
        sum(shift_requests[n][d - 1][s][k] * shifts[(n, d, s, k)] for n in all_nurses
            for d in all_days for s in all_shifts for k in all_nurse_per_shift))

    # Creates the solver and solve.
    solver = cp_model.CpSolver()
    status = solver.Solve(model)

    import openpyxl
    from openpyxl.styles import PatternFill

    # Create a new Excel workbook
    wb = openpyxl.Workbook()

    # Select the active sheet
    sheet = wb.active

    # red color for cell
    # red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    if status == cp_model.OPTIMAL:
        # sheet['B1'] = 'Morning (08.00 - 16.00)'
        # sheet['C1'] = 'Afternoon (16.00-24.00)'
        # sheet['D1'] = 'Night (24.00-8.00)'

        shift_cell_text = ['เช้า (08.00 - 16.00)', 'บ่าย (16.00-24.00)', 'ดึก (24.00-8.00)']
        print('Solution:')
        for d in all_days:
            print('Day', d, d in all_weekends)

            # cell_pos = 'A' + str(d + 2)
            # sheet[cell_pos] = d

            current_shift_cell = range((num_shifts * (d - 1)) + 1, (num_shifts * d) + 1)

            # Main hospital 'C' is the first person within shift, 'D' is the second person within shift, 'E' is the
            # third
            shift_column = ['C', 'D', 'E']

            sheet['A' + str(current_shift_cell[0])] = d

            for s in all_shifts:
                current_shift_column_count = 0
                sheet['B' + str(current_shift_cell[s])] = shift_cell_text[s]
                for k in all_nurse_per_shift:
                    for n in all_nurses:
                        if solver.Value(shifts[(n, d, s, k)]) == 1:
                            print(name_of_type_nurse(n), map_name_person(n), 'works shift', s, k)

                            if shift_requests[n][d - 1][s][k] == 1:

                                current_shift_column = shift_column[current_shift_column_count]

                                sheet[current_shift_column + str(current_shift_cell[s])] = map_name_person(
                                    n) + name_of_type_nurse(n)
                                current_shift_column_count += 1
                            else:
                                current_shift_column = shift_column[current_shift_column_count]

                                sheet[current_shift_column + str(current_shift_cell[s])] = map_name_person(
                                    n) + name_of_type_nurse(n)
                                current_shift_column_count += 1

            print()
        print(f'Number of shift requests met = {solver.ObjectiveValue()}',
              f'(out of {all_shifts})')

    else:
        print('No optimal solution found !')

    # Statistics.
    print('\nStatistics')
    print('  - conflicts: %i' % solver.NumConflicts())
    print('  - branches : %i' % solver.NumBranches())
    print('  - wall time: %f s' % solver.WallTime())

    morning_shifts = {}
    afternoon_shifts = {}
    night_shifts = {}

    morning_shifts_h = {}
    afternoon_shifts_h = {}
    night_shifts_h = {}

    morning_shifts_S = {}
    afternoon_shifts_S = {}
    night_shifts_S = {}

    morning_shifts_h_S = {}
    afternoon_shifts_h_S = {}
    night_shifts_h_S = {}

    special_shift = {}

    sum_all_shifts = 0

    for n in all_nurses:
        num_shifts_worked = 0
        num_shifts_holi = 0

        num_shifts_worked_S = 0
        num_shifts_holi_S = 0

        morning_shifts[n] = 0
        afternoon_shifts[n] = 0
        night_shifts[n] = 0

        morning_shifts_h[n] = 0
        afternoon_shifts_h[n] = 0
        night_shifts_h[n] = 0

        morning_shifts_S[n] = 0
        afternoon_shifts_S[n] = 0
        night_shifts_S[n] = 0

        morning_shifts_h_S[n] = 0
        afternoon_shifts_h_S[n] = 0
        night_shifts_h_S[n] = 0

        special_shift[n] = 0

        for d in all_days:
            for s in all_shifts:
                for k in all_nurse_per_shift:
                    if k != 2:
                        if d in all_weekends:
                            num_shifts_holi += solver.Value(shifts[(n, d, s, k)])
                            if s == 0:
                                morning_shifts_h[n] += solver.Value(shifts[(n, d, s, k)])
                            elif s == 1:
                                afternoon_shifts_h[n] += solver.Value(shifts[(n, d, s, k)])
                            elif s == 2:
                                night_shifts_h[n] += solver.Value(shifts[(n, d, s, k)])
                        else:
                            num_shifts_worked += solver.Value(shifts[(n, d, s, k)])
                            if s == 0:
                                morning_shifts[n] += solver.Value(shifts[(n, d, s, k)])
                            elif s == 1:
                                afternoon_shifts[n] += solver.Value(shifts[(n, d, s, k)])
                            elif s == 2:
                                night_shifts[n] += solver.Value(shifts[(n, d, s, k)])
                    else:
                        if d in all_weekends:
                            num_shifts_holi_S += solver.Value(shifts[(n, d, s, k)])
                            if s == 0:
                                morning_shifts_h_S[n] += solver.Value(shifts[(n, d, s, k)])
                            elif s == 1:
                                afternoon_shifts_h_S[n] += solver.Value(shifts[(n, d, s, k)])
                            elif s == 2:
                                night_shifts_h_S[n] += solver.Value(shifts[(n, d, s, k)])
                        else:
                            num_shifts_worked_S += solver.Value(shifts[(n, d, s, k)])
                            if s == 0:
                                morning_shifts_S[n] += solver.Value(shifts[(n, d, s, k)])
                            elif s == 1:
                                afternoon_shifts_S[n] += solver.Value(shifts[(n, d, s, k)])
                            elif s == 2:
                                night_shifts_S[n] += solver.Value(shifts[(n, d, s, k)])

        print(name_of_type_nurse(n), map_name_person(n), ' has ',
              num_shifts_worked, 'shifts normal day',
              num_shifts_holi, 'shifts holiday',
              num_shifts_worked + num_shifts_holi, 'total shifts')
        sum_all_shifts = sum_all_shifts + num_shifts_worked + num_shifts_holi

    # Print the results
    current_cell_num = [3, 22, 41, 60, 79]

    # Type of shift day
    sheet['I' + str(current_cell_num[0] - 2)] = 'วันทำการ'
    sheet['I' + str(current_cell_num[1] - 2)] = 'วันหยุด'
    sheet['I' + str(current_cell_num[2] - 2)] = 'วันทำการ รพ.เด็ก'
    sheet['I' + str(current_cell_num[3] - 2)] = 'วันหยุด รพ.เด็ก'
    sheet['I' + str(current_cell_num[4] - 2)] = 'สรุปผล'

    # Create Summary section
    shift_cell_text = ['เช้า (08.00 - 16.00)', 'บ่าย (16.00-24.00)', 'ดึก (24.00-8.00)', 'รวม']
    shift_cell_text_final_summary = ['วันหยุดชดเชย', 'วันหยุด', 'วันทำการ', 'รวม']
    field_summary = ['H', 'I', 'J', 'K']

    shift_cell_text_pos = 0
    for f in field_summary:
        sheet[f + str(current_cell_num[0] - 1)] = shift_cell_text[shift_cell_text_pos]
        sheet[f + str(current_cell_num[1] - 1)] = shift_cell_text[shift_cell_text_pos]
        sheet[f + str(current_cell_num[2] - 1)] = shift_cell_text[shift_cell_text_pos]
        sheet[f + str(current_cell_num[3] - 1)] = shift_cell_text[shift_cell_text_pos]
        sheet[f + str(current_cell_num[4] - 1)] = shift_cell_text_final_summary[shift_cell_text_pos]

        shift_cell_text_pos += 1

    for n in all_nurses:
        # Put name into summary cell (Summary section)
        sheet['G' + str(current_cell_num[0])] = map_name_person(n)
        sheet['G' + str(current_cell_num[1])] = map_name_person(n)
        sheet['G' + str(current_cell_num[2])] = map_name_person(n)
        sheet['G' + str(current_cell_num[3])] = map_name_person(n)
        sheet['G' + str(current_cell_num[4])] = map_name_person(n)

        # Put the value in record
        sum_all_holiday = solver.Value(morning_shifts_h[n]) + solver.Value(afternoon_shifts_h[n]) + solver.Value(
            night_shifts_h[n])

        sum_all_working_day = solver.Value(morning_shifts[n]) + solver.Value(afternoon_shifts[n]) + solver.Value(
            night_shifts[n])

        # Put the value in record
        sum_all_holiday_S = solver.Value(morning_shifts_h_S[n]) + solver.Value(afternoon_shifts_h_S[n]) + solver.Value(
            night_shifts_h_S[n])

        sum_all_working_day_S = solver.Value(morning_shifts_S[n]) + solver.Value(afternoon_shifts_S[n]) + solver.Value(
            night_shifts_S[n])

        # Each shift value in record
        shift_cell_record_val = [solver.Value(morning_shifts[n]), solver.Value(afternoon_shifts[n]),
                                 solver.Value(night_shifts[n]), sum_all_working_day]

        shift_cell_record_val_h = [solver.Value(morning_shifts_h[n]), solver.Value(afternoon_shifts_h[n]),
                                   solver.Value(night_shifts_h[n]), sum_all_holiday]

        # Each shift value in record
        shift_cell_record_val_S = [solver.Value(morning_shifts_S[n]), solver.Value(afternoon_shifts_S[n]),
                                   solver.Value(night_shifts_S[n]), sum_all_working_day_S]

        shift_cell_record_val_h_S = [solver.Value(morning_shifts_h_S[n]), solver.Value(afternoon_shifts_h_S[n]),
                                     solver.Value(night_shifts_h_S[n]), sum_all_holiday_S]

        shift_cell_final_record_val = [0, sum_all_holiday + sum_all_holiday_S,
                                       sum_all_working_day + sum_all_working_day_S,
                                       sum_all_holiday + sum_all_working_day + sum_all_holiday_S + sum_all_working_day_S]

        # Loop through all shifts
        shift_cell_text_pos = 0
        for f in field_summary:
            sheet[f + str(current_cell_num[0])] = shift_cell_record_val[shift_cell_text_pos]
            sheet[f + str(current_cell_num[1])] = shift_cell_record_val_h[shift_cell_text_pos]
            sheet[f + str(current_cell_num[2])] = shift_cell_record_val_S[shift_cell_text_pos]
            sheet[f + str(current_cell_num[3])] = shift_cell_record_val_h_S[shift_cell_text_pos]
            sheet[f + str(current_cell_num[4])] = shift_cell_final_record_val[shift_cell_text_pos]
            shift_cell_text_pos += 1

        current_cell_num[0] += 1
        current_cell_num[1] += 1
        current_cell_num[2] += 1
        current_cell_num[3] += 1
        current_cell_num[4] += 1

        print(map_name_person(n), " morning_shifts: ", solver.Value(morning_shifts[n]),
              " afternoon_shifts: ",
              solver.Value(afternoon_shifts[n]), " night_shifts: ", solver.Value(night_shifts[n]))

        print(map_name_person(n), " morning_shifts_h: ", solver.Value(morning_shifts_h[n]),
              " afternoon_shifts_h: ",
              solver.Value(afternoon_shifts_h[n]), " night_shifts_h: ", solver.Value(night_shifts_h[n]))

        print(" special_shifts: ", solver.Value(special_shift[n]))

        print('---')

    print('min working shift', min_morning_shift_per_nurse, min_afternoon_shift_per_nurse, min_night_shift_per_nurse)
    print('max working shift', max_morning_shift_per_nurse, max_afternoon_shift_per_nurse, max_night_shift_per_nurse)
    print('min holiday shift', min_morning_shift_per_nurse_h, min_afternoon_shift_per_nurse_h,
          min_night_shift_per_nurse_h)
    print('max holiday shift', max_morning_shift_per_nurse_h, max_afternoon_shift_per_nurse_h,
          max_night_shift_per_nurse_h)
    print(sum_all_shifts)
    print('sum min total shift', total_min)
    print('sum max total shift', total_max)

    # Save the workbook
    wb.save("sample.xlsx")


if __name__ == '__main__':
    main()
