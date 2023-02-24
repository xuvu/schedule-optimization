"""Nurse scheduling problem with shift requests."""
import math
import random

from ortools.sat.python import cp_model
import calendar
import numpy as np
from pandas import date_range
import itertools

stop_flag = [False, False]

name_of_person = ['ภูวเนตร',
                  'ราเชนทร์',
                  'พลกฤต',
                  'ชานนท์',
                  'ปรมะ',
                  'สัญญา',
                  'วินัย',
                  'รณยุทธ',
                  'วัฒพงษ์',
                  'ราเชน',
                  'นฤชิต',
                  'จีรวัฒน์',
                  'สุเมธร์',
                  'นที',
                  ]

# Type of nurses
# programmer = 8, service = 6
roles = ["Programmer", "Programmer", "Programmer", "Programmer", "Programmer", "Programmer", "Programmer", "Programmer",
         "Service", "Service", "Service", "Service", "Service", "Service"]

person_colors = ['FFC000',
                 'C2D69B',
                 '548DD4',
                 'CC26A7',
                 'FFFF00',
                 'C00000',
                 '92D050',
                 '00B0F0',
                 'D99594',
                 '938953',
                 '7030A0',
                 'FF7C80',
                 '00B050',
                 '00FFCC'
                 ]


# random colors
def random_colors(n):
    colors = []
    while len(colors) < n:
        red = random.randint(0, 255)
        green = random.randint(0, 255)
        blue = random.randint(0, 255)
        color = '{:02X}{:02X}{:02X}'.format(red, green, blue)
        if color not in colors:
            colors.append(color)
    return colors


# map person to color
def map_color_person(n):
    return person_colors[n]


# Get amount of each type
def get_amount_of_each_type(r):
    counts = {}
    for role in r:
        if role in counts:
            counts[role] += 1
        else:
            counts[role] = 1

    return counts


# Get the name of a nurse
def map_name_person(n):
    return name_of_person[n]


def weekend_days_in_month(year, month, additional_days):
    cal = calendar.Calendar()
    weekend_days = [day for day in cal.itermonthdays(year, month) if
                    day and calendar.weekday(year, month, day) >= 5]
    weekend_days.extend(additional_days)
    weekend_days = list(set(weekend_days))
    return weekend_days


def get_days_in_month(year, month):
    start_date = f'{year}-{month}-01'
    last_day = calendar.monthrange(year, month)[1]
    end_date = f'{year}-{month}-{last_day}'
    days = date_range(start=start_date, end=end_date, freq='D')
    day_list = [day.day for day in days]
    return day_list


def create_list(number_of_nurse, days_in_month, num_shift_per_day, num_nurse_per_shift):
    # create list which has shape like (number_of_nurse,days_in_month ,3)
    # 3 is for 3 shifts
    # 0 is for day off
    # 1 is for day on
    # 2 is for weekend
    return np.zeros((number_of_nurse, days_in_month, num_shift_per_day, num_nurse_per_shift))


def is_holiday(day, list_of_holiday):
    if day in list_of_holiday:
        return True
    else:
        return False


def get_max_diff_type(r, total_shifts):
    role_count = get_amount_of_each_type(r)

    min_value = min(role_count.values())
    min_type = [key for key, value in role_count.items() if value == min_value]

    max_value = max(role_count.values())
    max_type = [key for key, value in role_count.items() if value == max_value]

    list1 = [f"{k}{i}" for i, k in enumerate(r, 1) if k == str(max_type[0])]
    list2 = [f"{k}{i}" for i, k in enumerate(r, 1) if k == str(min_type[0])]

    min_shift = math.ceil(total_shifts / (len(list1) + len(list2)))

    element_counts = {element: 0 for element in list1 + list2}
    result = []

    count_pair = 0
    for k in range(math.ceil(total_shifts / (len(list1)) * len(list2))):
        for b in list2:
            for a in list1:
                if count_pair >= total_shifts:
                    break
                if b != a and element_counts[b] < min_shift and element_counts[a] < min_shift:
                    count_pair += 1
                    element_counts[a] += 1
                    element_counts[b] += 1
                    result += [(a, b)]
    return count_pair


# Define the types of nurses
types = ["Programmer", "Service"]
all_types = range(len(types))


# person_colors = random_colors(len(name_of_person))


# (EASY) (FINISH)
# no morning shift in working day => good

# (EASY) (FINISH)
# holiday can have all three shifts => good

# (FINISH)
# Each shift has 2 person

# (MEDIUM) (FINISH)
# 2 person per shift
# programmer + service => programmer + programmer => service + service

# (HARD) (FINISH)
# (1,2) same => (1,2) different => (0,1) same => good

# (HARD)
# night to morning is forbidden (0,2) in a day (can but shouldn't) => good

# (HARD)
# morning to night is forbidden (2,0) in a day (forbidden) => good

# special holiday

# special hospital
# afternoon to night (1,2) the highest priority with same person => (1,2) different person
# (0,1) if (1,2) can't happen


# Get type index of a nurse
def type_of_nurse(n):
    return types.index(roles[n])


# Get name of type of nurse
def name_of_type_nurse(n):
    return roles[n]


def main(m, decrement_):
    # This program tries to find an optimal assignment of nurses to shifts
    # (3 shifts per day, for 7 days), subject to some constraints (see below).
    # Each nurse can request to be assigned to specific shifts.
    # The optimal assignment maximizes the number of fulfilled shift requests

    year_ = 2023
    month_ = m

    total_min_working_all = 0
    total_max_working_all = 0
    total_min_holiday_all = 0
    total_max_holiday_all = 0
    total_min_summary = 0
    total_max_summary = 0
    total_min_main_hospital = 0
    total_max_main_hospital = 0
    total_min_child_hospital = 0
    total_max_child_hospital = 0

    num_nurses = len(roles)
    all_nurses = range(num_nurses)
    all_days = get_days_in_month(year_, month_)

    # example Add additional holiday 3,4,5
    # all_weekends = weekend_days_in_month(year_, month_, [3,4,5])
    all_weekends = weekend_days_in_month(year_, month_, [2,3])
    all_working_days = set(all_days) - set(all_weekends)

    total_working_days = len(all_working_days)
    total_holidays = len(all_weekends)

    # shift_requests[0][0][0] means first nurse, first day, first shift shift_requests[1][2][1] would mean that the
    # second nurse wants to work on the third day of the week on the second shift of the day.

    # Creates the model.
    model = cp_model.CpModel()

    # Creates shift variables.
    # Shifts[(n, d, s)]: nurse 'n' works shift 's' on day 'd'.
    shifts = {}
    num_shifts = 3
    all_shifts = range(num_shifts)  # [1,2,3,4]

    forbidden_shifts = [0]

    # Number of nurse per shift
    num_of_nurse_per_shift = 3
    all_nurse_per_shift = range(num_of_nurse_per_shift)

    # Amount of all shifts
    all_shift_count = (((total_holidays * (num_shifts - 0)) + (
            total_working_days * (num_shifts - len(forbidden_shifts)))) * num_of_nurse_per_shift)

    max_consecutive_shift = all_shift_count - decrement_['con_shift']
    # max_consecutive_shift = 100

    # reduce amount of all shift for satisfying the constraint
    # all_shift_count = all_shift_count - decrement_['diff_type']

    # this variable should be generated automatically then can be modified later on
    shift_requests = create_list(num_nurses, len(all_days), num_shifts, num_of_nurse_per_shift)  # <-- demo

    for n in all_nurses:
        for d in all_days:
            for s in all_shifts:
                for k in all_nurse_per_shift:
                    shifts[(n, d, s, k)] = model.NewBoolVar('shift_n%id%is%ik%i' % (n, d, s, k))

    """
    # Each shift is assigned to exactly one nurse in.
    for d in all_days:
        for s in all_shifts:
            model.Add(sum(shifts[(n, d, s)] for n in all_nurses) == 1)
    """

    # forbidden morning shift in working day
    for n in all_nurses:
        for d in all_working_days:
            for s in all_shifts:
                if s in forbidden_shifts:
                    for k in all_nurse_per_shift:
                        model.Add(shifts[(n, d, s, k)] == 0)

    # Each shift is assigned to exactly 1 nurse in.
    for d in all_days:
        if d not in all_weekends:  # not in weekend => in working day
            for s in all_shifts:
                for k in all_nurse_per_shift:
                    if s not in forbidden_shifts:  # exclude morning shift
                        model.Add(sum(shifts[(n, d, s, k)] for n in all_nurses) == 1)
        else:
            for s in all_shifts:
                for k in all_nurse_per_shift:
                    model.Add(sum(shifts[(n, d, s, k)] for n in all_nurses) == 1)

    # Each nurse works at most two shifts per day.
    maximum_each_day_shifts = 2
    for n in all_nurses:
        for d in all_days:
            shifts_per_day = sum(shifts[(n, d, s, k)] for s in all_shifts for k in all_nurse_per_shift)
            model.Add(shifts_per_day <= maximum_each_day_shifts)

    # Nurse type priority
    type_priority = [['Type1', 'Type2'], ['Type1', 'Type1'], ['Type2', 'Type2']]

    # total possible way of pairing ['Type1', 'Type2'] is the product of the number of people for each type

    # To ensure that each value of k0, k1 and k2 for every 's' cannot be assigned to the same nurse
    for d in all_days:
        for n in all_nurses:
            for s in all_shifts:
                model.Add(sum(shifts[(n, d, s, k)] for k in all_nurse_per_shift) <= 1)

    # Prioritize shift patterns (1,2) same
    max_consecutive_shift_count = 0

    day_interval = 1
    maximum_shifts = 2
    for d in all_days:
        for n in all_nurses:
            # For each nurse shouldn't have more than 2 shifts within 1+1 days
            if (d + day_interval) < len(all_days):  # 1 + (4-1) < 31
                model.Add(sum(shifts[(n, d + i, s, k)] for i in range(day_interval + 1) for s in all_shifts for k in all_nurse_per_shift) <= maximum_shifts)
            # shifts[(0, 1 + 0, 0, k), shifts[(0, 1 + 0, 1, k), shifts[(0, 1 + 0, 2, k)
            # shifts[(0, 1 + 1, 0, k), shifts[(0, 1 + 1, 1, k), shifts[(0, 1 + 1, 2, k)
            # shifts[(0, 1 + 2, 0, k), shifts[(0, 1 + 1, 1, k), shifts[(0, 1 + 2, 2, k) <= 2
            # all of the shifts period the nurse can only have 2 shifts maximum

    # Prioritize shift patterns (1,2) same
    for d in all_days:
        for n in all_nurses:
            if max_consecutive_shift_count < max_consecutive_shift:
                for k in all_nurse_per_shift:
                    # Desire shift pattern (1,2)
                    model.Add(shifts[(n, d, 1, k)] <= shifts[(n, d, 2, k)])  # (1,2)

                # Forbidden shifts pattern [(d,0),(d+1,0)]
                if d < len(all_days) - 1:
                    model.Add(sum(shifts[(n, d, 0, k)] for k in all_nurse_per_shift) + sum(shifts[(n, d + 1, 0, k)] for k in all_nurse_per_shift) <= 1)
            else:
                # Forbidden shifts pattern [0,2]
                model.Add(sum(shifts[(n, d, 0, k)] for k in all_nurse_per_shift) + sum(shifts[(n, d, 2, k)] for k in all_nurse_per_shift) <= 1)

                # Forbidden shifts pattern [(d,2),(d + 1,0)]
                if d < len(all_days) - 1:
                    model.Add(sum(shifts[(n, d, 2, k)] for k in all_nurse_per_shift) + sum(shifts[(n, d + 1, 0, k)] for k in all_nurse_per_shift) <= 1)
        max_consecutive_shift_count += 6

    # Forbidden shifts pattern [0,2] d,[(1,2,3)] d+1,[(1,2,3)]
    for n in all_nurses:
        for d in all_days:
            model.Add(sum(shifts[(n, d, 0, k)] for k in all_nurse_per_shift) + sum(
                shifts[(n, d, 2, k)] for k in all_nurse_per_shift) <= 1)
            if d < len(all_days) - 1:
                model.Add(sum(shifts[(n, d, 0, k)] for k in all_nurse_per_shift) + sum(
                    shifts[(n, d + 1, 0, k)] for k in all_nurse_per_shift) <= 1)

    # Each shift ensure different type of nurses
    max_diff = all_shift_count - decrement_['diff_type']
    max_diff_count = 0
    if stop_flag[0]:
        # Each shift ensure different type of nurses
        for d in all_weekends:
            for s in all_shifts:
                if max_diff_count <= max_diff:
                    type_count = [0] * len(all_types)
                    for k in range(0, num_of_nurse_per_shift - 1):  # -1 for excluding the third person
                        for t in all_types:
                            for n in all_nurses:
                                if type_of_nurse(n) == t:
                                    type_count[t] += shifts[(n, d, s, k)]
                    for t in all_types:
                        model.Add(type_count[t] <= 1)
                        max_diff_count += 1

        for d in all_working_days:
            for s in all_shifts:
                if s != 0:  # exclude morning and special hospital
                    if max_diff_count <= max_diff:
                        type_count = [0] * len(all_types)
                        for k in range(0, num_of_nurse_per_shift - 1):  # -1 for excluding the third person
                            for t in all_types:
                                for n in all_nurses:
                                    if type_of_nurse(n) == t:
                                        type_count[t] += shifts[(n, d, s, k)]
                        for t in all_types:
                            model.Add(type_count[t] <= 1)
                            max_diff_count += 1

    # All working shifts min-max variable
    # ---------------------------------------------------------------
    min_morning_shift_per_nurse = (total_working_days * (num_of_nurse_per_shift - 1)) // num_nurses
    min_afternoon_shift_per_nurse = (total_working_days * (num_of_nurse_per_shift - 1)) // num_nurses
    min_night_shift_per_nurse = (total_working_days * (num_of_nurse_per_shift - 1)) // num_nurses

    min_morning_shift_per_nurse_S = (total_working_days * (num_of_nurse_per_shift - 2)) // num_nurses
    min_afternoon_shift_per_nurse_S = (total_working_days * (num_of_nurse_per_shift - 2)) // num_nurses
    min_night_shift_per_nurse_S = (total_working_days * (num_of_nurse_per_shift - 2)) // num_nurses

    if min_morning_shift_per_nurse_S == 0 and (total_working_days * (num_of_nurse_per_shift - 2)) != 0:
        min_morning_shift_per_nurse_S = 0
        max_morning_shift_per_nurse_S = 1

        min_afternoon_shift_per_nurse_S = 0
        max_afternoon_shift_per_nurse_S = 1

        min_night_shift_per_nurse_S = 0
        max_night_shift_per_nurse_S = 1


    elif (total_working_days * (num_of_nurse_per_shift - 2)) % num_nurses == 0:
        max_morning_shift_per_nurse_S = min_morning_shift_per_nurse_S
        max_afternoon_shift_per_nurse_S = min_morning_shift_per_nurse_S
        max_night_shift_per_nurse_S = min_morning_shift_per_nurse_S
    else:
        max_morning_shift_per_nurse_S = min_morning_shift_per_nurse_S + 1
        max_afternoon_shift_per_nurse_S = min_morning_shift_per_nurse_S + 1
        max_night_shift_per_nurse_S = min_morning_shift_per_nurse_S + 1

    if min_morning_shift_per_nurse == 0 and (total_working_days * (num_of_nurse_per_shift - 1)) != 0:
        min_morning_shift_per_nurse = 0
        max_morning_shift_per_nurse = 1

        min_afternoon_shift_per_nurse = 0
        max_afternoon_shift_per_nurse = 1

        min_night_shift_per_nurse = 0
        max_night_shift_per_nurse = 1

    elif (total_working_days * (num_of_nurse_per_shift - 1)) % num_nurses == 0:
        max_morning_shift_per_nurse = min_morning_shift_per_nurse
        max_afternoon_shift_per_nurse = min_morning_shift_per_nurse
        max_night_shift_per_nurse = min_morning_shift_per_nurse
    else:
        max_morning_shift_per_nurse = min_morning_shift_per_nurse + 1
        max_afternoon_shift_per_nurse = min_morning_shift_per_nurse + 1
        max_night_shift_per_nurse = min_morning_shift_per_nurse + 1
    # ---------------------------------------------------------------

    # All holiday shifts min-max variable
    # ---------------------------------------------------------------
    min_morning_shift_per_nurse_h = (total_holidays * (num_of_nurse_per_shift - 1)) // num_nurses
    min_afternoon_shift_per_nurse_h = (total_holidays * (num_of_nurse_per_shift - 1)) // num_nurses
    min_night_shift_per_nurse_h = (total_holidays * (num_of_nurse_per_shift - 1)) // num_nurses

    min_morning_shift_per_nurse_h_S = (total_holidays * (num_of_nurse_per_shift - 2)) // num_nurses
    min_afternoon_shift_per_nurse_h_S = (total_holidays * (num_of_nurse_per_shift - 2)) // num_nurses
    min_night_shift_per_nurse_h_S = (total_holidays * (num_of_nurse_per_shift - 2)) // num_nurses

    if min_morning_shift_per_nurse_h_S == 0 and (total_holidays * (num_of_nurse_per_shift - 2)) != 0:
        min_morning_shift_per_nurse_h_S = 0
        max_morning_shift_per_nurse_h_S = 1

        min_afternoon_shift_per_nurse_h_S = 0
        max_afternoon_shift_per_nurse_h_S = 1

        min_night_shift_per_nurse_h_S = 0
        max_night_shift_per_nurse_h_S = 1

    elif (total_holidays * (num_of_nurse_per_shift - 2)) % num_nurses == 0:
        max_morning_shift_per_nurse_h_S = min_morning_shift_per_nurse_h_S
        max_afternoon_shift_per_nurse_h_S = min_afternoon_shift_per_nurse_h_S
        max_night_shift_per_nurse_h_S = min_night_shift_per_nurse_h_S

    else:
        max_morning_shift_per_nurse_h_S = min_morning_shift_per_nurse_h_S + 1
        max_afternoon_shift_per_nurse_h_S = min_afternoon_shift_per_nurse_h_S + 1
        max_night_shift_per_nurse_h_S = min_night_shift_per_nurse_h_S + 1

    if min_morning_shift_per_nurse_h == 0 and (total_holidays * (num_of_nurse_per_shift - 1)) != 0:
        min_morning_shift_per_nurse_h = 0
        max_morning_shift_per_nurse_h = 1

        min_afternoon_shift_per_nurse_h = 0
        max_afternoon_shift_per_nurse_h = 1

        min_night_shift_per_nurse_h = 0
        max_night_shift_per_nurse_h = 1


    elif (total_holidays * (num_of_nurse_per_shift - 1)) % num_nurses == 0:
        max_morning_shift_per_nurse_h = min_morning_shift_per_nurse_h
        max_afternoon_shift_per_nurse_h = min_afternoon_shift_per_nurse_h
        max_night_shift_per_nurse_h = min_night_shift_per_nurse_h
    else:
        max_morning_shift_per_nurse_h = min_morning_shift_per_nurse_h + 1
        max_afternoon_shift_per_nurse_h = min_afternoon_shift_per_nurse_h + 1
        max_night_shift_per_nurse_h = min_night_shift_per_nurse_h + 1
    # ---------------------------------------------------------------

    # Distribute all day for main hospital
    for n in all_nurses:
        num_shift_worked = 0
        num_shift_worked_h = 0

        num_shifts_morning = 0
        num_shifts_morning_h = 0

        num_shifts_afternoon = 0
        num_shifts_afternoon_h = 0

        num_shifts_night = 0
        num_shifts_night_h = 0

        num_shift_worked_S = 0
        num_shift_worked_h_S = 0

        num_shifts_morning_S = 0
        num_shifts_morning_h_S = 0

        num_shifts_afternoon_S = 0
        num_shifts_afternoon_h_S = 0

        num_shifts_night_S = 0
        num_shifts_night_h_S = 0

        for k in all_nurse_per_shift:
            if k != 2:
                # Distribute working day shifts
                # ---------------------------------------------------------------
                for d in all_working_days:
                    for s in all_shifts:
                        if s == 1:
                            num_shift_worked += shifts[(n, d, s, k)]
                            num_shifts_afternoon += shifts[(n, d, s, k)]

                        elif s == 2:
                            num_shift_worked += shifts[(n, d, s, k)]
                            num_shifts_night += shifts[(n, d, s, k)]

                # Distribute holiday day shifts
                # ---------------------------------------------------------------
                for d in all_weekends:
                    for s in all_shifts:
                        if s == 0:
                            num_shift_worked_h += shifts[(n, d, s, k)]
                            num_shifts_morning_h += shifts[(n, d, s, k)]
                        elif s == 1:
                            num_shift_worked_h += shifts[(n, d, s, k)]
                            num_shifts_afternoon_h += shifts[(n, d, s, k)]
                        elif s == 2:
                            num_shift_worked_h += shifts[(n, d, s, k)]
                            num_shifts_night_h += shifts[(n, d, s, k)]
            else:
                # Distribute working day shifts (S)
                # ---------------------------------------------------------------
                for d in all_working_days:
                    for s in all_shifts:
                        if s == 1:
                            num_shift_worked_S += shifts[(n, d, s, k)]
                            num_shifts_afternoon_S += shifts[(n, d, s, k)]

                        elif s == 2:
                            num_shift_worked_S += shifts[(n, d, s, k)]
                            num_shifts_night_S += shifts[(n, d, s, k)]

                # Distribute holiday day shifts (S)
                # ---------------------------------------------------------------
                for d in all_weekends:
                    for s in all_shifts:
                        if s == 0:
                            num_shift_worked_h_S += shifts[(n, d, s, k)]
                            num_shifts_morning_h_S += shifts[(n, d, s, k)]
                        elif s == 1:
                            num_shift_worked_h_S += shifts[(n, d, s, k)]
                            num_shifts_afternoon_h_S += shifts[(n, d, s, k)]
                        elif s == 2:
                            num_shift_worked_h_S += shifts[(n, d, s, k)]
                            num_shifts_night_h_S += shifts[(n, d, s, k)]

        # # all shifts constraint working day
        # model.Add(min_shifts_per_nurse <= num_shift_worked)
        # model.Add(num_shift_worked <= max_shifts_per_nurse + 1)
        #
        # # all shifts constraint holiday
        # model.Add(min_shifts_per_nurse_h <= num_shift_worked_h)
        # model.Add(num_shift_worked_h <= max_shifts_per_nurse_h)
        #
        # # all shifts constraint working day (S)
        # model.Add(min_shifts_per_nurse_S-1 <= num_shift_worked_S)
        # model.Add(num_shift_worked_S <= max_shifts_per_nurse_S)
        #
        # # all shifts constraint holiday (S)
        # model.Add(min_shifts_per_nurse_h_S <= num_shift_worked_h_S)
        # model.Add(num_shift_worked_h_S <= max_shifts_per_nurse_h_S + 1)

        # morning shifts constraint (Forbidden)
        # model.Add(min_morning_shift_per_nurse <= num_shifts_morning)
        # model.Add(num_shifts_morning <= max_morning_shift_per_nurse)

        # conflict between
        # minimum type within each shift
        # distribution constraint

        # soften the constraint of working day shifts
        slack = 0

        # afternoon shifts constraint working day
        model.Add(min_afternoon_shift_per_nurse - slack <= num_shifts_afternoon)
        model.Add(num_shifts_afternoon <= max_afternoon_shift_per_nurse + slack)

        # night shifts constraint working day
        model.Add(min_night_shift_per_nurse - slack <= num_shifts_night)
        model.Add(num_shifts_night <= max_night_shift_per_nurse + slack)

        # afternoon shifts constraint working day (S)
        model.Add(min_afternoon_shift_per_nurse_S - slack <= num_shifts_afternoon_S)
        model.Add(num_shifts_afternoon_S <= max_afternoon_shift_per_nurse_S + slack)

        # night shifts constraint working day (S)
        model.Add(min_night_shift_per_nurse_S - slack <= num_shifts_night_S)
        model.Add(num_shifts_night_S <= max_night_shift_per_nurse_S + slack)

        # soften the constraint for holiday shifts
        slack_h = 0

        # morning shifts constraint holiday
        model.Add(min_morning_shift_per_nurse_h - slack_h <= num_shifts_morning_h)
        model.Add(num_shifts_morning_h <= max_morning_shift_per_nurse_h + slack_h)

        # afternoon shifts constraint holiday
        model.Add(min_afternoon_shift_per_nurse_h - slack_h <= num_shifts_afternoon_h)
        model.Add(num_shifts_afternoon_h <= max_afternoon_shift_per_nurse_h + slack_h)

        # night shifts constraint holiday
        model.Add(min_night_shift_per_nurse_h - slack_h <= num_shifts_night_h)
        model.Add(num_shifts_night_h <= max_night_shift_per_nurse_h + slack_h)

        # morning shifts constraint holiday (S)
        model.Add(min_morning_shift_per_nurse_h_S - slack_h <= num_shifts_morning_h_S)
        model.Add(num_shifts_morning_h_S <= max_morning_shift_per_nurse_h_S + slack_h)

        # afternoon shifts constraint holiday (S)
        model.Add(min_afternoon_shift_per_nurse_h_S - slack_h <= num_shifts_afternoon_h_S)
        model.Add(num_shifts_afternoon_h_S <= max_afternoon_shift_per_nurse_h_S + slack_h)

        # night shifts constraint holiday (S)
        model.Add(min_night_shift_per_nurse_h_S - slack_h <= num_shifts_night_h_S)
        model.Add(num_shifts_night_h_S <= max_night_shift_per_nurse_h_S + slack_h)

        # total min-max distribution sum of (main hospital)
        # num_of_nurse_per_shift - 1 remove k2
        max_slack = 0
        min_slack = -1
        total_max_main_hospital = (((total_holidays * (num_shifts - 0)) + (
                total_working_days * (num_shifts - len(forbidden_shifts)))) * (num_of_nurse_per_shift - 1))
        total_max_main_hospital = math.ceil(total_max_main_hospital / num_nurses) + max_slack
        total_min_main_hospital = total_max_main_hospital - 1 - max_slack + min_slack

        # Distribute sum of (main hospital)
        model.Add(total_min_main_hospital <= num_shift_worked + num_shift_worked_h)
        model.Add(num_shift_worked + num_shift_worked_h <= total_max_main_hospital)

        # total min-max distribution sum of (child hospital)
        # num_of_nurse_per_shift - 1 remove k0 and k1
        max_slack = 0
        min_slack = -1
        total_max_child_hospital = (((total_holidays * (num_shifts - 0)) + (
                total_working_days * (num_shifts - len(forbidden_shifts)))) * (num_of_nurse_per_shift - 2))
        total_max_child_hospital = math.ceil(total_max_child_hospital / num_nurses) + max_slack
        total_min_child_hospital = total_max_child_hospital - 1 - max_slack + min_slack

        # Distribute sum of (child hospital)
        model.Add(total_min_child_hospital <= num_shift_worked_S + num_shift_worked_h_S)
        model.Add(num_shift_worked_S + num_shift_worked_h_S <= total_max_child_hospital)

        # total min-max distribution sum of all shift (holiday summary)
        max_slack = 0
        min_slack = 0
        total_max_holiday_all = total_holidays * (num_of_nurse_per_shift * num_shifts)
        total_max_holiday_all = math.ceil(total_max_holiday_all / num_nurses) + max_slack
        total_min_holiday_all = total_max_holiday_all - 1 - max_slack + min_slack

        # Distribute sum of all shift (holiday summary)
        model.Add(total_min_holiday_all <= num_shift_worked_h + num_shift_worked_h_S)
        model.Add(num_shift_worked_h + num_shift_worked_h_S <= total_max_holiday_all)

        # total min-max distribution sum of all shift (morning_holiday summary)
        max_slack = 0
        min_slack = 0
        total_max_morning_holiday_all = total_holidays * (num_shifts-2) * num_of_nurse_per_shift
        total_max_morning_holiday_all = math.ceil(total_max_morning_holiday_all / num_nurses) + max_slack
        total_min_morning_holiday_all = total_max_morning_holiday_all - 1 - max_slack + min_slack

        # Distribute sum of all shift (holiday summary)
        model.Add(total_min_morning_holiday_all <= num_shifts_morning_h_S + num_shifts_morning_h)
        model.Add(num_shifts_morning_h_S + num_shifts_morning_h <= total_max_morning_holiday_all)

        # total min-max distribution sum of all shift (working summary)
        max_slack = 0
        min_slack = -1
        total_max_working_all = total_working_days * (num_of_nurse_per_shift * (num_shifts - len(forbidden_shifts)))
        total_max_working_all = math.ceil(total_max_working_all / num_nurses) + max_slack
        total_min_working_all = total_max_working_all - 1 - max_slack + min_slack

        # Distribute sum of all shift (working + holiday summary)
        model.Add(total_min_working_all <= num_shift_worked + num_shift_worked_S)
        model.Add(num_shift_worked + num_shift_worked_S <= total_max_working_all)

        # total min-max distribution sum of all shift (Summary)
        max_slack = 0
        min_slack = 0
        total_max_summary = math.ceil(all_shift_count / num_nurses) + max_slack
        total_min_summary = total_max_summary - 1 - max_slack + min_slack

        # Distribute sum of all shift (Summary)
        model.Add(
            total_min_summary <= num_shift_worked_S + num_shift_worked + num_shift_worked_h_S + num_shift_worked_h)
        model.Add(
            num_shift_worked_S + num_shift_worked + num_shift_worked_h_S + num_shift_worked_h <= total_max_summary)
        # ---------------------------------------------------------------

    print('morning range (working day)', min_morning_shift_per_nurse, '<=', max_morning_shift_per_nurse)
    print('afternoon range (working day)', min_afternoon_shift_per_nurse, '<=', max_afternoon_shift_per_nurse)
    print('night range (working day)', min_night_shift_per_nurse, '<=', max_night_shift_per_nurse)
    print('---')
    print('morning range (holiday)', min_morning_shift_per_nurse_h, '<=', max_morning_shift_per_nurse_h)
    print('afternoon range (holiday)', min_afternoon_shift_per_nurse_h, '<=', max_afternoon_shift_per_nurse_h)
    print('night range (holiday)', min_night_shift_per_nurse_h, '<=', max_night_shift_per_nurse_h)
    print('---')
    print('morning range (working day)(S)', min_morning_shift_per_nurse_S, '<=', max_morning_shift_per_nurse_S)
    print('afternoon range (working day)(S)', min_afternoon_shift_per_nurse_S, '<=', max_afternoon_shift_per_nurse_S)
    print('night range (working day)(S)', min_night_shift_per_nurse_S, '<=', max_night_shift_per_nurse_S)
    print('---')
    print('morning range (holiday)(S)', min_morning_shift_per_nurse_h_S, '<=', max_morning_shift_per_nurse_h_S)
    print('afternoon range (holiday)(S)', min_afternoon_shift_per_nurse_h_S, '<=', max_afternoon_shift_per_nurse_h_S)
    print('night range (holiday)(S)', min_night_shift_per_nurse_h_S, '<=', max_night_shift_per_nurse_h_S)
    print('---')
    print('Total range main hospital', total_min_main_hospital, '<=', total_max_main_hospital)
    print('Total range child hospital', total_min_child_hospital, '<=', total_max_child_hospital)
    print('---')
    print('Total summary range (working day)', total_min_working_all, '<=', total_max_working_all)
    print('Total summary range (holiday)', total_min_holiday_all, '<=', total_max_holiday_all)
    print('Total summary range', total_min_summary, '<=', total_max_summary)

    print('max-con-shifts', max_consecutive_shift)
    print('max-diff-shifts', max_diff)
    print('current month', m)

    # pylint: disable=g-complex-comprehension
    model.Maximize(
        sum(shift_requests[n][d - 1][s][k] * shifts[(n, d, s, k)] for n in all_nurses
            for d in all_days for s in all_shifts for k in all_nurse_per_shift)
    )

    # Creates the solver and solve.
    solver = cp_model.CpSolver()
    status = solver.Solve(model)

    import openpyxl
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Border, Side

    # Create a new Excel workbook
    wb = openpyxl.Workbook()

    # Select the active sheet
    sheet = wb.active

    redFill = PatternFill(start_color='FF6B6B',
                          end_color='FF6B6B',
                          fill_type='solid')
    blueFill = PatternFill(start_color='ADD8E6',
                           end_color='ADD8E6',
                           fill_type='solid')
    greyFill = PatternFill(start_color='D3D3D3',
                           end_color='D3D3D3',
                           fill_type='solid')

    # red color for cell
    # red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    if status == cp_model.OPTIMAL:
        if stop_flag[0]:
            stop_flag[1] = True
        else:
            stop_flag[0] = True

        # sheet['B1'] = 'Morning (08.00 - 16.00)'
        # sheet['C1'] = 'Afternoon (16.00-24.00)'
        # sheet['D1'] = 'Night (24.00-8.00)'

        shift_cell_text = ['08:00-16:00', '16:00-24:00', '24:00-08:00']

        # Main hospital 'C' is the first person within shift, 'D' is the second person within shift, 'E' is the
        # third
        shift_column = [['C', 'D', 'E'], ['H', 'I', 'J']]
        time_day_column = [['A', 'B'], ['F', 'G']]

        current_shift_column = shift_column[0]
        current_time_day_column = time_day_column[0]

        print('Solution:')
        count_person_cell = [[[], [], []], [[], [], []]]
        for d in all_days:
            print('Day', d, d in all_weekends)

            # List of days within the month
            current_shift_cell = range((num_shifts * (d - 1)) + 1, (num_shifts * d) + 1)

            # List of days within the month
            current_shift_cell_ = range((num_shifts * (d - 1)) + 97, (num_shifts * d) + 97)

            if d > len(all_days) // 2:
                current_shift_column = shift_column[1]
                current_time_day_column = time_day_column[1]

                # List of days within the month
                current_shift_cell = range((num_shifts * ((d - len(all_days) // 2) - 1)) + 1,
                                           ((num_shifts * (d - len(all_days) // 2)) + 1))

                # List of days within the month
                current_shift_cell_ = range((num_shifts * ((d - len(all_days) // 2) - 1)) + 97,
                                            ((num_shifts * (d - len(all_days) // 2)) + 97))

                # color for weekend
            if d not in all_weekends:
                sheet[current_time_day_column[0] + str(current_shift_cell[0])] = str(d)
                sheet[current_time_day_column[0] + str(current_shift_cell_[0])] = str(d)
            else:
                sheet[current_time_day_column[0] + str(current_shift_cell[0])] = str(d) + str('*')
                sheet[current_time_day_column[0] + str(current_shift_cell_[0])] = str(d) + str('*')

                sheet[current_time_day_column[0] + str(current_shift_cell[0])].fill = redFill
                sheet[current_time_day_column[0] + str(current_shift_cell_[0])].fill = redFill

                sheet[current_time_day_column[0] + str(current_shift_cell[0] + 1)].fill = redFill
                sheet[current_time_day_column[0] + str(current_shift_cell_[0] + 1)].fill = redFill

                sheet[current_time_day_column[0] + str(current_shift_cell[0] + 2)].fill = redFill
                sheet[current_time_day_column[0] + str(current_shift_cell_[0] + 2)].fill = redFill

            for s in all_shifts:
                if s != 0 and d in all_working_days:  # skip morning shift
                    sheet[current_time_day_column[1] + str(current_shift_cell[s])] = shift_cell_text[s]
                    sheet[current_time_day_column[1] + str(current_shift_cell_[s])] = shift_cell_text[s]
                elif d in all_weekends:
                    sheet[current_time_day_column[1] + str(current_shift_cell[s])] = shift_cell_text[s]
                    sheet[current_time_day_column[1] + str(current_shift_cell_[s])] = shift_cell_text[s]

                    sheet[current_time_day_column[1] + str(current_shift_cell[s])].fill = redFill
                    sheet[current_time_day_column[1] + str(current_shift_cell_[s])].fill = redFill

                pair_val = []
                temp_count_person_cell = []
                for k in all_nurse_per_shift:
                    for n in all_nurses:
                        if solver.Value(shifts[(n, d, s, k)]) == 1:
                            print(name_of_type_nurse(n), '(' + map_name_person(n) + ')', 'works shift', s, k)
                            pair_val += [(k, n)]

                            # if shift_requests[n][d - 1][s][k] == 1:
                            # if type_of_nurse(n) == 0 and k == 1:
                            #     sheet[shift_column[0] + str(current_shift_cell[s])].fill = blueFill
                            #     sheet[shift_column[0] + str(current_shift_cell[s])] = map_name_person(n)
                            # elif type_of_nurse(n) == 1 and k == 1:
                            #     sheet[shift_column[0] + str(current_shift_cell[s])].fill = greyFill
                            #     sheet[shift_column[0] + str(current_shift_cell[s])] = map_name_person(n)

                if len(pair_val) != 0 and type_of_nurse(pair_val[1][1]) == 0:
                    temp_pair_val = pair_val[1]
                    pair_val[1] = pair_val[0]
                    pair_val[0] = temp_pair_val
                print(pair_val)

                column_count = 0
                for pair_ in pair_val:
                    current_color = PatternFill(start_color=map_color_person(pair_[1]),
                                                end_color=map_color_person(pair_[1]),
                                                fill_type='solid')
                    if type_of_nurse(pair_[1]) == 0:
                        sheet[current_shift_column[column_count] + str(current_shift_cell[s])].fill = blueFill
                        sheet[current_shift_column[column_count] + str(current_shift_cell[s])] = map_name_person(
                            pair_[1])

                        sheet[current_shift_column[column_count] + str(current_shift_cell_[s])].fill = current_color
                        sheet[current_shift_column[column_count] + str(current_shift_cell_[s])] = map_name_person(
                            pair_[1])

                        temp_count_person_cell += [
                            '$' + current_shift_column[column_count] + '$' + str(current_shift_cell_[s])]
                        column_count += 1
                    else:
                        sheet[current_shift_column[column_count] + str(current_shift_cell[s])].fill = greyFill
                        sheet[current_shift_column[column_count] + str(current_shift_cell[s])] = map_name_person(
                            pair_[1])

                        sheet[current_shift_column[column_count] + str(current_shift_cell_[s])].fill = current_color
                        sheet[current_shift_column[column_count] + str(current_shift_cell_[s])] = map_name_person(
                            pair_[1])

                        temp_count_person_cell += [
                            '$' + current_shift_column[column_count] + '$' + str(current_shift_cell_[s])]
                        column_count += 1

                if d not in all_weekends:
                    # for working day
                    count_person_cell[0][s] += [temp_count_person_cell]
                else:
                    # for holiday
                    count_person_cell[1][s] += [temp_count_person_cell]

        print(f'Number of shift requests met = {solver.ObjectiveValue()}', f'(out of {all_shifts})')

        # Statistics.
        print('\nStatistics')
        print('  - conflicts: %i' % solver.NumConflicts())
        print('  - branches : %i' % solver.NumBranches())
        print('  - wall time: %f s' % solver.WallTime())

        morning_shifts = {}
        afternoon_shifts = {}
        night_shifts = {}

        morning_shifts_h = {}
        afternoon_shifts_h = {}
        night_shifts_h = {}

        morning_shifts_S = {}
        afternoon_shifts_S = {}
        night_shifts_S = {}

        morning_shifts_h_S = {}
        afternoon_shifts_h_S = {}
        night_shifts_h_S = {}

        special_shift = {}

        sum_all_shifts = 0

        for n in all_nurses:
            num_shifts_worked = 0
            num_shifts_holi = 0

            num_shifts_worked_S = 0
            num_shifts_holi_S = 0

            morning_shifts[n] = 0
            afternoon_shifts[n] = 0
            night_shifts[n] = 0

            morning_shifts_h[n] = 0
            afternoon_shifts_h[n] = 0
            night_shifts_h[n] = 0

            morning_shifts_S[n] = 0
            afternoon_shifts_S[n] = 0
            night_shifts_S[n] = 0

            morning_shifts_h_S[n] = 0
            afternoon_shifts_h_S[n] = 0
            night_shifts_h_S[n] = 0

            special_shift[n] = 0

            for d in all_days:
                for s in all_shifts:
                    for k in all_nurse_per_shift:
                        if k != 2:
                            if d in all_weekends:
                                num_shifts_holi += solver.Value(shifts[(n, d, s, k)])
                                if s == 0:
                                    morning_shifts_h[n] += solver.Value(shifts[(n, d, s, k)])
                                elif s == 1:
                                    afternoon_shifts_h[n] += solver.Value(shifts[(n, d, s, k)])
                                elif s == 2:
                                    night_shifts_h[n] += solver.Value(shifts[(n, d, s, k)])
                            else:
                                num_shifts_worked += solver.Value(shifts[(n, d, s, k)])
                                if s == 0:
                                    morning_shifts[n] += solver.Value(shifts[(n, d, s, k)])
                                elif s == 1:
                                    afternoon_shifts[n] += solver.Value(shifts[(n, d, s, k)])
                                elif s == 2:
                                    night_shifts[n] += solver.Value(shifts[(n, d, s, k)])
                        else:
                            if d in all_weekends:
                                num_shifts_holi_S += solver.Value(shifts[(n, d, s, k)])
                                if s == 0:
                                    morning_shifts_h_S[n] += solver.Value(shifts[(n, d, s, k)])
                                elif s == 1:
                                    afternoon_shifts_h_S[n] += solver.Value(shifts[(n, d, s, k)])
                                elif s == 2:
                                    night_shifts_h_S[n] += solver.Value(shifts[(n, d, s, k)])
                            else:
                                num_shifts_worked_S += solver.Value(shifts[(n, d, s, k)])
                                if s == 0:
                                    morning_shifts_S[n] += solver.Value(shifts[(n, d, s, k)])
                                elif s == 1:
                                    afternoon_shifts_S[n] += solver.Value(shifts[(n, d, s, k)])
                                elif s == 2:
                                    night_shifts_S[n] += solver.Value(shifts[(n, d, s, k)])

            print(name_of_type_nurse(n), map_name_person(n), ' has ',
                  num_shifts_worked + num_shifts_worked_S, 'shifts normal day',
                  num_shifts_holi + num_shifts_holi_S, 'shifts holiday',
                  num_shifts_worked + num_shifts_holi + num_shifts_worked_S + num_shifts_holi_S, 'total shifts')
            sum_all_shifts = sum_all_shifts + num_shifts_worked + num_shifts_holi + num_shifts_worked_S + num_shifts_holi_S

        # Print the results
        current_cell_num = [3, 6 + num_nurses]
        summary_cell_num = (num_nurses * 2) + 9

        # Create Summary section
        shift_cell_text = ['08:00-16:00', '16:00-24:00', '24:00-08:00', 'รวม']
        shift_cell_text_final_summary = ['วันหยุดเช้า', 'วันหยุด', 'วันทำการ', 'รวม']
        field_name = [['M', 'N', 'O', 'P'], ['S', 'T', 'U', 'V'], ['Y', 'Z', 'AA', 'AB']]
        fixable_section_cell_num = 97

        for n in range(len(all_nurses)):
            current_color = PatternFill(start_color=map_color_person(n),
                                        end_color=map_color_person(n),
                                        fill_type='solid')
            for count in range(4):
                temp = []
                temp_ = []
                for f in range(len(field_name)):
                    temp += [field_name[f][count] + str(current_cell_num[0] + fixable_section_cell_num + n)]
                    temp_ += [field_name[f][count] + str(current_cell_num[1] + fixable_section_cell_num + n)]
                sheet[temp[2]] = '=SUM(' + temp[0] + ',' + temp[1] + ')'
                sheet[temp[2]].fill = current_color

                sheet[temp_[2]] = '=SUM(' + temp_[0] + ',' + temp_[1] + ')'
                sheet[temp_[2]].fill = current_color

        # field_name = [['M', 'N', 'O', 'P'], ['S', 'T', 'U', 'V'], ['X', 'Y', 'Z', 'AB']]
        header_field = [field_name[0][2], field_name[1][2], field_name[2][2]]
        # for main hospital header
        sheet[header_field[0] + str(current_cell_num[0] - 2)] = 'วันทำการ'
        sheet[header_field[1] + str(current_cell_num[0] - 2)] = 'วันหยุด'
        sheet[header_field[2] + str(current_cell_num[0] - 2)] = 'รวมโรงพยาบาลหลัก'

        # for child hospital header
        sheet[header_field[0] + str(current_cell_num[1] - 2)] = 'วันทำการ โรงพยาบาลเด็ก'
        sheet[header_field[1] + str(current_cell_num[1] - 2)] = 'วันหยุด โรงพยาบาลเด็ก'
        sheet[header_field[2] + str(current_cell_num[1] - 2)] = 'รวมโรงพยาบาลเด็ก'

        # for summary header
        sheet[header_field[0] + str(summary_cell_num - 1)] = 'สรุป'

        # for main hospital header
        sheet[header_field[0] + str(current_cell_num[0] - 2 + fixable_section_cell_num)] = 'วันทำการ'
        sheet[header_field[1] + str(current_cell_num[0] - 2 + fixable_section_cell_num)] = 'วันหยุด'
        sheet[header_field[2] + str(current_cell_num[0] - 2 + fixable_section_cell_num)] = 'รวมโรงพยาบาลหลัก'

        # for child hospital header
        sheet[header_field[0] + str(current_cell_num[1] - 2 + fixable_section_cell_num)] = 'วันทำการ โรงพยาบาลเด็ก'
        sheet[header_field[1] + str(current_cell_num[1] - 2 + fixable_section_cell_num)] = 'วันหยุด โรงพยาบาลเด็ก'
        sheet[header_field[2] + str(current_cell_num[1] - 2 + fixable_section_cell_num)] = 'รวมโรงพยาบาลเด็ก'

        # for summary header
        sheet[header_field[0] + str(summary_cell_num - 1 + fixable_section_cell_num)] = 'สรุป'

        for c in current_cell_num:
            for j in range(len(field_name)):
                shift_cell_text_pos = 0
                for k in range(len(field_name[j])):

                    # Non-Editable section in Excel
                    sheet[field_name[j][k] + str(c - 1)] = shift_cell_text[shift_cell_text_pos]
                    sheet[field_name[j][k] + str(c - 1)] = shift_cell_text[shift_cell_text_pos]
                    sheet[field_name[j][k] + str(c - 1)] = shift_cell_text[shift_cell_text_pos]
                    sheet[field_name[j][k] + str(c - 1)] = shift_cell_text[shift_cell_text_pos]

                    # Editable section in Excel
                    sheet[field_name[j][k] + str(c - 1 + fixable_section_cell_num)] = shift_cell_text[shift_cell_text_pos]
                    sheet[field_name[j][k] + str(c - 1 + fixable_section_cell_num)] = shift_cell_text[shift_cell_text_pos]
                    sheet[field_name[j][k] + str(c - 1 + fixable_section_cell_num)] = shift_cell_text[shift_cell_text_pos]
                    sheet[field_name[j][k] + str(c - 1 + fixable_section_cell_num)] = shift_cell_text[shift_cell_text_pos]

                    shift_cell_text_pos += 1

            shift_cell_text_pos = 0
            for k in range(len(field_name[0])):
                sheet[field_name[0][k] + str(summary_cell_num)] = shift_cell_text_final_summary[shift_cell_text_pos]
                sheet[field_name[0][k] + str(summary_cell_num + fixable_section_cell_num)] = shift_cell_text_final_summary[shift_cell_text_pos]
                shift_cell_text_pos += 1

        name_field = ['L', 'R', 'X']

        # Put the Value in the cell for each field
        for n in all_nurses:

            current_color = PatternFill(start_color=map_color_person(n),
                                        end_color=map_color_person(n),
                                        fill_type='solid')

            # Detail name
            for c in range(len(current_cell_num)):
                for nf in name_field:
                    if type_of_nurse(n) == 0:
                        sheet[nf + str(current_cell_num[c])].fill = blueFill
                        sheet[nf + str(current_cell_num[c])] = map_name_person(n)

                        sheet[nf + str(current_cell_num[c] + fixable_section_cell_num)].fill = current_color
                        sheet[nf + str(current_cell_num[c] + fixable_section_cell_num)] = map_name_person(n)
                    else:
                        sheet[nf + str(current_cell_num[c])].fill = greyFill
                        sheet[nf + str(current_cell_num[c])] = map_name_person(n)

                        sheet[nf + str(current_cell_num[c] + fixable_section_cell_num)].fill = current_color
                        sheet[nf + str(current_cell_num[c] + fixable_section_cell_num)] = map_name_person(n)

            # Summary name
            # field_name = [['M', 'N', 'O', 'P'], ['T', 'U', 'V', 'W'], ['Z', 'AA', 'AB', 'AC']]
            if type_of_nurse(n) == 0:
                sheet[name_field[0] + str(summary_cell_num + 1)].fill = blueFill
                sheet[name_field[0] + str(summary_cell_num + 1)] = map_name_person(n)

                sheet[name_field[0] + str(summary_cell_num + 1 + fixable_section_cell_num)].fill = current_color
                sheet[name_field[0] + str(summary_cell_num + 1 + fixable_section_cell_num)] = map_name_person(n)
            else:
                sheet[name_field[0] + str(summary_cell_num + 1)].fill = greyFill
                sheet[name_field[0] + str(summary_cell_num + 1)] = map_name_person(n)

                sheet[name_field[0] + str(summary_cell_num + 1 + fixable_section_cell_num)].fill = current_color
                sheet[name_field[0] + str(summary_cell_num + 1 + fixable_section_cell_num)] = map_name_person(n)

            # sum of all shifts for main hospital in holiday
            sum_all_holiday = solver.Value(morning_shifts_h[n]) + solver.Value(afternoon_shifts_h[n]) + solver.Value(
                night_shifts_h[n])

            # sum of all shifts for main hospital in working day
            sum_all_working_day = solver.Value(morning_shifts[n]) + solver.Value(afternoon_shifts[n]) + solver.Value(
                night_shifts[n])

            # sum of all shifts for child hospital in holiday
            sum_all_holiday_S = solver.Value(morning_shifts_h_S[n]) + solver.Value(
                afternoon_shifts_h_S[n]) + solver.Value(night_shifts_h_S[n])

            # sum of all shifts for child hospital in working day
            sum_all_working_day_S = solver.Value(morning_shifts_S[n]) + solver.Value(afternoon_shifts_S[n]) + solver.Value(night_shifts_S[n])

            # Each shift value in record of main hospital in working day
            shift_cell_record_val = [solver.Value(morning_shifts[n]), solver.Value(afternoon_shifts[n]),
                                     solver.Value(night_shifts[n]), sum_all_working_day]

            # Each shift value in record of main hospital in holiday
            shift_cell_record_val_h = [solver.Value(morning_shifts_h[n]), solver.Value(afternoon_shifts_h[n]),
                                       solver.Value(night_shifts_h[n]), sum_all_holiday]

            # Each shift value in record of child hospital in working day
            shift_cell_record_val_S = [solver.Value(morning_shifts_S[n]), solver.Value(afternoon_shifts_S[n]),
                                       solver.Value(night_shifts_S[n]), sum_all_working_day_S]

            # Each shift value in record of child hospital in holiday
            shift_cell_record_val_h_S = [solver.Value(morning_shifts_h_S[n]), solver.Value(afternoon_shifts_h_S[n]),
                                         solver.Value(night_shifts_h_S[n]), sum_all_holiday_S]

            # summary of all shifts in main hospital
            sum_main_hospital_shifts = [solver.Value(morning_shifts[n]) + solver.Value(morning_shifts_h[n]),
                                        solver.Value(afternoon_shifts[n]) + solver.Value(afternoon_shifts_h[n]),
                                        solver.Value(night_shifts[n]) + solver.Value(night_shifts_h[n]),
                                        sum_all_working_day + sum_all_holiday]

            # summary of all shifts in child hospital
            sum_child_hospital_shifts = [solver.Value(morning_shifts_S[n]) + solver.Value(morning_shifts_h_S[n]),
                                         solver.Value(afternoon_shifts_S[n]) + solver.Value(afternoon_shifts_h_S[n]),
                                         solver.Value(night_shifts_S[n]) + solver.Value(night_shifts_h_S[n]),
                                         sum_all_working_day_S + sum_all_holiday_S]
            # Final summary
            shift_cell_record_summary = [solver.Value(morning_shifts_h_S[n]) + solver.Value(morning_shifts_h[n]), sum_all_holiday + sum_all_holiday_S,
                                         sum_all_working_day + sum_all_working_day_S,
                                         sum_all_holiday + sum_all_working_day + sum_all_holiday_S + sum_all_working_day_S]

            count_cell_name = ['morning', 'afternoon', 'night']
            count_cell_working_main = {'morning': '', 'afternoon': '', 'night': ''}
            count_cell_working_child = {'morning': '', 'afternoon': '', 'night': ''}
            for s in range(len(count_person_cell[0])):
                for k in range(len(count_person_cell[0][s])):
                    start_end = []  # for main hospital
                    temp = []  # for child hospital
                    if len(count_person_cell[0][s][k]) > 0:
                        for j in range(len(count_person_cell[0][s][k])):
                            if j != 2:  # Exclude child hospital
                                start_end += [count_person_cell[0][s][k][j]]  # add start and end cell for counting
                            else:
                                temp += [count_person_cell[0][s][k][j]]  # for child hospital
                        if k != 0:
                            count_cell_working_main[count_cell_name[s]] += '+COUNTIF(' + start_end[0] + ':' + start_end[1] + ',' + '"' + map_name_person(n) + '")'  # add start and end cell for counting
                            count_cell_working_child[count_cell_name[s]] += '+COUNTIF(' + temp[0] + ',' + '"' + map_name_person(n) + '")'
                        else:
                            count_cell_working_main[count_cell_name[s]] += '=COUNTIF(' + start_end[0] + ':' + start_end[1] + ',' + '"' + map_name_person(n) + '")'  # add start and end cell for counting
                            count_cell_working_child[count_cell_name[s]] += '=COUNTIF(' + temp[0] + ',' + '"' + map_name_person(n) + '")'

            count_cell_holiday_main = {'morning': '', 'afternoon': '', 'night': ''}
            count_cell_holiday_child = {'morning': '', 'afternoon': '', 'night': ''}
            for s in range(len(count_person_cell[1])):
                for k in range(len(count_person_cell[1][s])):
                    start_end = []  # for main hospital
                    temp = []  # for child hospital
                    if len(count_person_cell[1][s][k]) > 0:
                        for j in range(len(count_person_cell[1][s][k])):
                            if j != 2:  # Exclude child hospital
                                start_end += [count_person_cell[1][s][k][j]]  # add start and end cell for counting
                            else:
                                temp += [count_person_cell[1][s][k][j]]  # for child hospital
                        if k != 0:
                            count_cell_holiday_main[count_cell_name[s]] += '+COUNTIF(' + start_end[0] + ':' + start_end[1] + ',' + '"' + map_name_person(n) + '")'  # add start and end cell for counting
                            count_cell_holiday_child[count_cell_name[s]] += '+COUNTIF(' + temp[0] + ',' + '"' + map_name_person(n) + '")'
                        else:
                            count_cell_holiday_main[count_cell_name[s]] += '=COUNTIF(' + start_end[0] + ':' + start_end[1] + ',' + '"' + map_name_person(n) + '")'  # add start and end cell for counting
                            count_cell_holiday_child[count_cell_name[s]] += '=COUNTIF(' + temp[0] + ',' + '"' + map_name_person(n) + '")'

            # Main hospital summary
            main_h = [shift_cell_record_val, shift_cell_record_val_h, sum_main_hospital_shifts]

            # Main hospital cell count
            main_count = [
                [count_cell_working_main['morning'], count_cell_working_main['afternoon'],
                 count_cell_working_main['night']],
                [count_cell_holiday_main['morning'], count_cell_holiday_main['afternoon'],
                 count_cell_holiday_main['night']]
            ]

            # Loop through all shifts
            for h in range(len(main_h)):
                for j in range(len(field_name)):
                    for k in range(len(field_name[j])):
                        sheet[field_name[j][k] + str(current_cell_num[0])] = main_h[j][k]

            # Loop through all shifts (cell count)
            for h in range(len(main_h)):
                for j in range(len(field_name) - 1):
                    sum_cell = []
                    for k in range(len(field_name[j]) - 1):
                        sheet[field_name[j][k] + str(current_cell_num[0] + fixable_section_cell_num)] = main_count[j][k]
                        sheet[field_name[j][k] + str(current_cell_num[0] + fixable_section_cell_num)].fill = current_color
                        sum_cell += [field_name[j][k] + str(current_cell_num[0] + fixable_section_cell_num)]

                    sheet[field_name[j][3] + str(current_cell_num[0] + fixable_section_cell_num)] = '=SUM(' + sum_cell[0] + ':' + sum_cell[len(sum_cell) - 1] + ')'
                    sheet[field_name[j][3] + str(current_cell_num[0] + fixable_section_cell_num)].fill = current_color

            # Child hospital summary
            child_h = [shift_cell_record_val_S, shift_cell_record_val_h_S, sum_child_hospital_shifts]

            # Child hospital cell count
            child_count = [
                [count_cell_working_child['morning'], count_cell_working_child['afternoon'],
                 count_cell_working_child['night']],
                [count_cell_holiday_child['morning'], count_cell_holiday_child['afternoon'],
                 count_cell_holiday_child['night']]
            ]

            # Loop through all shifts
            for h in range(len(main_h)):
                for j in range(len(field_name)):
                    for k in range(len(field_name[j])):
                        sheet[field_name[j][k] + str(current_cell_num[1])] = child_h[j][k]
            # Loop through all shifts (cell count)
            for h in range(len(main_h)):
                for j in range(len(field_name) - 1):
                    sum_cell = []
                    for k in range(len(field_name[j]) - 1):
                        sheet[field_name[j][k] + str(current_cell_num[1] + fixable_section_cell_num)] = child_count[j][k]
                        sheet[field_name[j][k] + str(current_cell_num[1] + fixable_section_cell_num)].fill = current_color
                        sum_cell += [field_name[j][k] + str(current_cell_num[1] + fixable_section_cell_num)]
                    sheet[field_name[j][3] + str(current_cell_num[1] + fixable_section_cell_num)] = '=SUM(' + sum_cell[0] + ':' + sum_cell[len(sum_cell) - 1] + ')'
                    sheet[field_name[j][3] + str(current_cell_num[1] + fixable_section_cell_num)].fill = current_color

            # Final Summary
            # final_field = ['AE', 'AF', 'AG', 'AH']
            for column_ in range(len(field_name[0])):
                sheet[str(field_name[0][column_]) + str(summary_cell_num + 1)] = shift_cell_record_summary[column_]

            # Final Summary fixable for working day
            sum_working_day_cell = field_name[0][len(field_name[0]) - 2] + str(summary_cell_num + 1 + fixable_section_cell_num)
            working_day_main = str(field_name[0][len(field_name[0]) - 1]) + str(current_cell_num[0] + fixable_section_cell_num)
            working_day_child = str(field_name[0][len(field_name[0]) - 1]) + str(current_cell_num[1] + fixable_section_cell_num)
            sheet[sum_working_day_cell] = '=SUM(' + working_day_main + ',' + working_day_child + ')'
            sheet[sum_working_day_cell].fill = current_color

            # Final Summary fixable for holiday
            sum_holiday_cell = field_name[0][len(field_name[0]) - 3] + str(summary_cell_num + 1 + fixable_section_cell_num)
            holiday_main = str(field_name[1][len(field_name[0]) - 1]) + str(current_cell_num[0] + fixable_section_cell_num)
            holiday_child = str(field_name[1][len(field_name[0]) - 1]) + str(current_cell_num[1] + fixable_section_cell_num)
            sheet[sum_holiday_cell] = '=SUM(' + holiday_main + ',' + holiday_child + ')'
            sheet[sum_holiday_cell].fill = current_color

            # Final Summary  fixable for morning holiday
            sum_morning_holiday_cell = field_name[0][len(field_name[0]) - 4] + str(summary_cell_num + 1 + fixable_section_cell_num)
            morning_holiday_main = str(field_name[1][len(field_name[0]) - 4]) + str(current_cell_num[0] + fixable_section_cell_num)
            morning_holiday_child = str(field_name[1][len(field_name[0]) - 4]) + str(current_cell_num[1] + fixable_section_cell_num)
            sheet[sum_morning_holiday_cell] = '=SUM(' + morning_holiday_main + ',' + morning_holiday_child + ')'
            sheet[sum_morning_holiday_cell].fill = current_color

            # Final Summary working + holiday
            sum_all_cell = field_name[0][len(field_name[0]) - 1] + str(summary_cell_num + 1 + fixable_section_cell_num)
            sheet[sum_all_cell] = '=SUM(' + sum_working_day_cell + ',' + sum_holiday_cell + ')'
            sheet[sum_all_cell].fill = current_color

            print(map_name_person(n), " morning_shifts: ", solver.Value(morning_shifts[n]),
                  " afternoon_shifts: ",
                  solver.Value(afternoon_shifts[n]), " night_shifts: ", solver.Value(night_shifts[n]))

            print(map_name_person(n), " morning_shifts_h: ", solver.Value(morning_shifts_h[n]),
                  " afternoon_shifts_h: ",
                  solver.Value(afternoon_shifts_h[n]), " night_shifts_h: ", solver.Value(night_shifts_h[n]))

            print(map_name_person(n), " morning_shifts (S): ", solver.Value(morning_shifts_S[n]),
                  " afternoon_shifts (S): ",
                  solver.Value(afternoon_shifts_S[n]), " night_shifts (S): ", solver.Value(night_shifts_S[n]))

            print(map_name_person(n), " morning_shifts_h (S): ", solver.Value(morning_shifts_h_S[n]),
                  " afternoon_shifts_h (S): ",
                  solver.Value(afternoon_shifts_h_S[n]), " night_shifts_h (S): ", solver.Value(night_shifts_h_S[n]))

            print('---')

            current_cell_num[0] += 1
            current_cell_num[1] += 1
            summary_cell_num += 1

        print(sum_all_shifts)
        # Save the workbook
        if stop_flag[1]:
            # Add borders to cells containing text or numbers
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
                for cell in row:
                    if cell.value is not None:
                        cell.border = thin_border

            wb.save("sample" + str(m) + ".xlsx")
        return True
    else:
        print('No optimal solution found !')
        return False


if __name__ == '__main__':
    all_month = range(3, 4)
    decrement = {'diff_type': 0, 'con_shift': -6}
    for m in all_month:
        while True:
            if not stop_flag[0]:
                decrement['con_shift'] += 6
                main(m, decrement)
            else:
                main(m, decrement)
                decrement['diff_type'] += 2
                if stop_flag[1]:
                    stop_flag[0] = False
                    stop_flag[1] = False
                    decrement['diff_type'] = 0
                    decrement['con_shift'] = -6
                    break
