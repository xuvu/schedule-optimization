import openpyxl
from openpyxl.styles import PatternFill

# Create a new Excel workbook
wb = openpyxl.Workbook()

# Select the active sheet
sheet = wb.active

blueFill = PatternFill(start_color='ADD8E6',
                           end_color='ADD8E6',
                           fill_type='solid')

sheet['A10'].fill = blueFill
sheet['A10'] = 10

wb.save("colorFill.xlsx")